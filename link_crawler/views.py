from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from rest_framework import viewsets
from django.contrib import messages
from .models import Link, Domain_Blogger_Details
from .serializers import LinkSerializer
import datetime
from django.http import HttpResponse, HttpResponseRedirect
from django.views.decorators.http import require_http_methods
from .models import Link
from django.utils.dateparse import parse_date
import openpyxl
from datetime import datetime
from django.contrib.messages import get_messages
from django.conf import settings
import os
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse
from django.utils.timezone import now
from .tasks import check_selected_urls_index, send_email




@login_required
def home(request):
    # Start with the base queryset
    links_queryset = Link.objects.filter(user=request.user).order_by('-last_index_check', '-link_created')    
    # Retrieve filter/search parameters
    start_date = request.GET.get('startDate', '').strip()
    end_date = request.GET.get('endDate', '').strip()
    link_type = request.GET.get('linkType', '').strip()
    index_status = request.GET.get('index_status', '').strip()
    print("Index status:",index_status)

    search_target_link = request.GET.get('searchTargetLink', '').strip()

    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            links_queryset = links_queryset.filter(link_created__range=[start_date, end_date])
        except ValueError:
            # Handle invalid date format
            pass

    # Apply link type filter if provided and not the placeholder
    if link_type and link_type != "Choose...":
        if link_type == "Others":
            links_queryset = links_queryset.exclude(status_of_link__in=["Dofollow", "Nofollow", "404", "Link Removed"])
        else:
            links_queryset = links_queryset.filter(status_of_link=link_type)

    # Apply index status filter if provided and not the placeholder
    if index_status and index_status != "Choose...":
        if index_status == "Others":
            links_queryset = links_queryset.exclude(index_status__in=["Index", "Not Index"])
        else:
            links_queryset = links_queryset.filter(index_status=index_status)

    # Apply target link search if provided
    if search_target_link:
        links_queryset = links_queryset.filter(target_link__icontains=search_target_link)
        
    total_links = links_queryset.count()
    # Get 'per_page' from GET parameters or default to 20, and make sure it's an integer
    per_page = int(request.GET.get('per_page', 20))

    # Instantiate the Paginator
    paginator = Paginator(links_queryset, per_page)

    # Get the page number from the request
    page = request.GET.get('page', 1)
    try:
        links = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver the first page
        links = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver the last page
        links = paginator.page(paginator.num_pages)

    return render(request, 'link_crawler/home.html', {
        'links': links,
        'per_page': per_page,
        'current_page': int(page),
        'total_links': total_links,
        'start_date': start_date,
        'end_date': end_date,
        'link_type': link_type,
        'index_status': index_status,
        'search_target_link': search_target_link,
    })
    
    
@login_required
def handle_actions(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        print(action)
        selected_ids = request.POST.getlist('selected_links')

        if action == 'download':
            # Assuming download_status_report returns an HttpResponse object
            return download_status_report(request, selected_ids)
        elif action == 'delete':
            # Assuming delete_links performs deletion and then returns an HttpResponse object
            return delete_links(request, selected_ids)
        elif action == 'mark_addressed':
            return mark_links_as_addressed(request, selected_ids)        
        elif action == 'check_indexation':
            # Trigger the background task for checking indexation of selected links
            check_selected_urls_index.delay(selected_ids)
            # Redirect or respond to indicate the task is underway
            return HttpResponseRedirect(reverse('home'))
        elif action == "send_email":
            send_email.delay(selected_ids)
             # Redirect or respond to indicate the task is underway
            return HttpResponseRedirect(reverse('home'))
        else:
            # If the action is not recognized, redirect to a default page or show an error message
            return HttpResponseRedirect(reverse('home'))

    # If the request method is not POST, redirect to a default page or show an error message
    return HttpResponseRedirect(reverse('home'))

def download_status_report(request, selected_links_ids):
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define the title of the worksheet
    ws.title = "Links Status Report"

    # Set the header row
    headers = ['Target Link', 'Referring Domain', 'Anchor', 'Status Of Link', 'Is Index', 'Link Created', 'Last Crawled']
    ws.append(headers)
    # Fetch the selected links from the database
    selected_links = Link.objects.filter(id__in=selected_links_ids)

    # Iterate over the queryset and append rows to the worksheet
    for link in selected_links:
        if link.index_status == Link.index:
            is_index = 'Yes' 
        elif link.index_status == Link.not_index:
            is_index = 'No' 
        else:
            is_index = 'Unknown' 
        # Prepare row with handling of potential blank values
        row = [
            link.target_link or '',  # Empty string if None
            link.link_to or '',  # Empty string if None
            link.anchor_text or '',  # Empty string if None
            link.get_status_of_link_display() or '',  # Empty string if None
            is_index,
            link.link_created.strftime('%Y-%m-%d') if link.link_created else '',
            link.last_crawl_date.strftime('%Y-%m-%d') if link.last_crawl_date else '',
        ]
        ws.append(row)

    # Set the HTTP response with a file attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=links_status_report.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

def delete_links(request, selected_links_ids):
    links_to_delete = Link.objects.filter(id__in=selected_links_ids)
    count = links_to_delete.count()
    links_to_delete.delete()

    messages.success(request, f'Deleted {count} links successfully.')
    return HttpResponseRedirect(reverse('home'))

def mark_links_as_addressed(request, selected_links_ids):
    updated_count = Link.objects.filter(id__in=selected_links_ids).update(address_status='addressed')

    # Provide feedback to the user
    messages.success(request, f'Marked {updated_count} links as addressed successfully.')

    # Redirect to home or another success page
    return HttpResponseRedirect(reverse('home'))


class LinkViewSet(viewsets.ModelViewSet):
    queryset = Link.objects.all()
    serializer_class = LinkSerializer


@login_required
def download_report(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Links Report"

    headers = ['Target Link', 'Link To', 'Anchor Text', 'Status Of Link', 'Index Status', 'Link Created', 'Last Crawl Date']
    ws.append(headers)

    status_filter = request.GET.get('report_type', None)
    print("Requested Status Filter:", status_filter)  # Debug print

    if status_filter:
        if status_filter in ["indexed", "not_indexed"]:
            links = Link.objects.filter(index_status=status_filter)
        else:
            links = Link.objects.filter(status_of_link=status_filter)
    else:
        links = Link.objects.all()

    print("Number of links found:", len(links))  # Debug print

    for link in links:
        data_row = [
            link.target_link,
            link.link_to,
            link.anchor_text,
            link.get_status_of_link_display(),
            link.get_index_status_display(),
            link.link_created.strftime('%Y-%m-%d') if link.link_created else 'N/A',
            link.last_crawl_date.strftime('%Y-%m-%d') if link.last_crawl_date else 'N/A',
        ]
        ws.append(data_row)

    filename = f"Links_Report_{now().strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response




@require_http_methods(["GET", "POST"])
@login_required  # Require the user to be logged in to access this view
def add_links(request):
    if request.method == 'POST':
        if 'fileUpload' not in request.FILES:
            messages.error(request, "No file uploaded.")
            return redirect('add_links')
        
        excel_file = request.FILES.get('fileUpload')
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "File is not in the format of a .xlsx")
            return redirect('add_links')

        skipped_rows = []

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                link_created = datetime(*row[4].timetuple()[:3]) if isinstance(row[4], datetime) else None

                # Check if a record with the same target_link, link_to, anchor_text, and user exists
                exists = Link.objects.filter(
                    target_link=row[2],
                    link_to=row[3],
                    anchor_text=row[1],
                    user=request.user  # Filter by the current user
                ).exists()

                if exists:
                    skipped_rows.append({
                        'row': index-1,
                        'anchor': row[1],
                        'target_link': row[2],
                        'link_to': row[3],
                    })
                else:
                    Link.objects.create(
                        user=request.user,  # Set the current user
                        anchor_text=row[1],
                        target_link=row[2],
                        link_to=row[3],
                        link_created=link_created
                    )

            if skipped_rows:
                request.session['skipped_rows'] = skipped_rows
                messages.warning(request, "Some rows were skipped due to duplicates.")
            else:
                messages.success(request, "Excel file processed successfully")

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

        return redirect('add_links')

    skipped_rows = request.session.pop('skipped_rows', None)
    return render(request, "link_crawler/add_links.html", {'skipped_rows': skipped_rows})

@login_required
def add_blogger_details(request):
    if request.method == 'POST':
        if 'fileUpload' not in request.FILES:
            messages.error(request, "No file uploaded.")
            return redirect('add_blogger_details')

        excel_file = request.FILES.get('fileUpload')
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "File is not in the format of a .xlsx")
            return redirect('add_blogger_details')

        skipped_rows = []
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                url = row[0]
                blogger_name = row[1]
                blogger_email = row[2]

                # Check if a record with the same url exists
                exists = Domain_Blogger_Details.objects.filter(url=url).exists()
                if exists:
                    skipped_rows.append({
                        'row': index - 1,
                        'url': url,
                        'blogger_name': blogger_name,
                        'blogger_email': blogger_email,
                    })
                else:
                    Domain_Blogger_Details.objects.create(
                        url=url,
                        blogger_name=blogger_name,
                        blogger_email=blogger_email
                    )

            if skipped_rows:
                request.session['skipped_rows'] = skipped_rows
                messages.warning(request, "Some rows were skipped due to duplicates.")
            else:
                messages.success(request, "Excel file processed successfully")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('add_blogger_details')

        skipped_rows = request.session.pop('skipped_rows', None)
        return render(request, "link_crawler/add_blogger_details.html", {'skipped_rows': skipped_rows})

    return render(request, "link_crawler/add_blogger_details.html")

@login_required
def download_excel_template(request):
    # Define the path to the Excel file
    file_path = os.path.join(settings.BASE_DIR, 'links_upload_format.xlsx')
    
    # Check if file exists
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="links_upload_format.xlsx"'
            return response
    else:
        # You can return an HTTP 404 response if the file is not found
        # Or handle it some other way if you prefer
        return HttpResponse("The requested Excel template was not found.", status=404)
    
@login_required
def download_excel_template_blogger(request):
    # Define the path to the Excel file
    file_path = os.path.join(settings.BASE_DIR, 'links_upload_format_blogger.xlsx')
    
    # Check if file exists
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="links_upload_format.xlsx"'
            return response
    else:
        # You can return an HTTP 404 response if the file is not found
        # Or handle it some other way if you prefer
        return HttpResponse("The requested Excel template was not found.", status=404)
